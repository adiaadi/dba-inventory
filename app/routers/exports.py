from datetime import UTC, datetime
from io import BytesIO
from typing import Iterable
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.core.config import get_settings
from app.db.session import get_db
from app.models import DatabaseInstance, Host
from app.routers.common import apply_database_filters, apply_host_filters
from app.routers.dashboard import (
    DB_TYPE_VIEWS,
    database_product_label,
    database_version_label,
    detected_db_type,
    detected_db_type_by_zabbix_rules,
    detected_server_platform_from_values,
    host_search_text,
    is_family_database_asset,
    is_zabbix_server_asset,
    normalized_db_type,
    normalized_virtual_filter,
    operating_system_label,
    server_core_label,
    server_model_label,
    server_ram_label,
    server_vendor_label,
    unique_hosts,
    virtual_status_label,
)
from app.services.zabbix_refresh import maybe_refresh_zabbix_cache

router = APIRouter(prefix="/exports", tags=["exports"])


def apply_sheet_style(ws, headers: Iterable[str]) -> None:
    header_fill = PatternFill("solid", fgColor="E9EEF5")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = max(14, len(header) + 2)


def workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def excel_cell_value(value):
    if not isinstance(value, datetime):
        return value

    settings = get_settings()
    try:
        timezone = ZoneInfo(settings.app_timezone)
    except ZoneInfoNotFoundError:
        timezone = UTC

    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(timezone).replace(tzinfo=None)


def append_excel_row(ws, values: Iterable) -> None:
    ws.append([excel_cell_value(value) for value in values])


def view_db_type(view: str | None, db_type: str | None) -> str | None:
    view_config = DB_TYPE_VIEWS.get(view or "")
    if view_config:
        return view_config["label"]
    if db_type:
        return normalized_db_type(db_type)
    return None


def export_db_type_label(value: str | None) -> str:
    if value == "SQLServer":
        return "SQLServer"
    return value or "Servers"


def export_filename(value: str, asset_label: str) -> str:
    safe_value = value.lower().replace(" ", "_")
    safe_asset = asset_label.lower().replace(" ", "_")
    return f"dba_inventory_{safe_value}_{safe_asset}.xlsx"


def dbms_database_export_rows(
    db: Session,
    family: str,
    environment: str | None = None,
    role: str | None = None,
    monitoring_status: str | None = None,
) -> list[dict]:
    stmt = select(Host).options(selectinload(Host.databases)).order_by(Host.hostname)
    stmt = apply_host_filters(stmt, None, environment, role, monitoring_status)
    hosts = unique_hosts(
        [host for host in db.scalars(stmt).all() if is_family_database_asset(host, family)]
    )

    rows = []
    for host in hosts:
        db_type = detected_db_type_by_zabbix_rules(host) or detected_db_type(host) or host.db_type or family
        rows.append(
            {
                "instance_name": host.zabbix_host_name or host.hostname,
                "ip": host.ip_address or "-",
                "environment": (host.environment or "-").upper(),
                "db_type": db_type,
                "product": database_product_label(host, family) or "-",
                "version": database_version_label(host, family) or "-",
            }
        )
    return rows


def server_summary_export_rows(
    db: Session,
    db_type: str | None = None,
    environment: str | None = None,
    role: str | None = None,
    monitoring_status: str | None = None,
    virtual: str | None = None,
    view: str | None = None,
) -> list[dict]:
    stmt = select(Host).options(selectinload(Host.databases)).order_by(Host.hostname)
    stmt = apply_host_filters(stmt, None, environment, role, monitoring_status)
    hosts = db.scalars(stmt).all()

    host_db_labels = {
        host.id: detected_db_type_by_zabbix_rules(host) or detected_db_type(host)
        for host in hosts
    }
    requested_db_type = view_db_type(view, db_type)
    if requested_db_type:
        hosts = [host for host in hosts if host_db_labels.get(host.id) == requested_db_type]

    server_hosts = unique_hosts([host for host in hosts if is_zabbix_server_asset(host)])
    host_model_labels = {host.id: server_model_label(host) for host in server_hosts}
    host_vendor_labels = {host.id: server_vendor_label(host) for host in server_hosts}
    host_platform_labels = {
        host.id: detected_server_platform_from_values(
            host_vendor_labels.get(host.id),
            host_model_labels.get(host.id),
            host_search_text(host),
        )
        for host in server_hosts
    }
    host_virtual_labels = {
        host.id: virtual_status_label(host_platform_labels.get(host.id))
        for host in server_hosts
    }

    active_virtual_filter = normalized_virtual_filter(virtual)
    if active_virtual_filter:
        server_hosts = [
            host
            for host in server_hosts
            if host_virtual_labels.get(host.id) == active_virtual_filter
        ]

    rows = []
    for host in server_hosts:
        instance_db_types = sorted({database.db_type for database in host.databases if database.db_type})
        rows.append(
            {
                "server": host.zabbix_host_name or host.hostname,
                "ip": host.ip_address or "-",
                "environment": (host.environment or "-").upper(),
                "db_type": host_db_labels.get(host.id) or host.db_type or ", ".join(instance_db_types) or "-",
                "virtual": host_virtual_labels.get(host.id) or "-",
                "server_model": host_model_labels.get(host.id) or "-",
                "server_vendor": host_vendor_labels.get(host.id) or "-",
                "core": server_core_label(host),
                "ram": server_ram_label(host),
                "operating_system": operating_system_label(host),
            }
        )
    return rows


@router.get("/hosts.xlsx")
def export_hosts(
    db_type: str | None = None,
    environment: str | None = None,
    role: str | None = None,
    monitoring_status: str | None = None,
    virtual: str | None = None,
    view: str | None = None,
    asset_view: str | None = None,
    db: Session = Depends(get_db),
):
    maybe_refresh_zabbix_cache(db)
    requested_db_type = view_db_type(view, db_type)
    if requested_db_type and asset_view == "databases":
        rows = dbms_database_export_rows(
            db,
            requested_db_type,
            environment=environment,
            role=role,
            monitoring_status=monitoring_status,
        )
        db_type_label = export_db_type_label(requested_db_type)
        headers = ["Instance Name", "IP", "Environment", "DB Type", "Product", "Version"]
        workbook = Workbook()
        ws = workbook.active
        ws.title = f"{db_type_label} Databases"
        ws.append(headers)
        for row in rows:
            append_excel_row(
                ws,
                [
                    row["instance_name"],
                    row["ip"],
                    row["environment"],
                    row["db_type"],
                    row["product"],
                    row["version"],
                ]
            )
        apply_sheet_style(ws, headers)
        return workbook_response(workbook, export_filename(db_type_label, "databases"))

    rows = server_summary_export_rows(
        db,
        db_type=db_type,
        environment=environment,
        role=role,
        monitoring_status=monitoring_status,
        virtual=virtual,
        view=view,
    )
    db_type_label = export_db_type_label(requested_db_type)
    filename = (
        export_filename(db_type_label, "servers")
        if requested_db_type
        else "dba_inventory_server_summary.xlsx"
    )

    headers = [
        "Server",
        "IP",
        "Environment",
        "DB Type",
        "Virtual",
        "Server model",
        "Server vendor",
        "Core",
        "RAM",
        "Operating system",
    ]
    workbook = Workbook()
    ws = workbook.active
    ws.title = f"{db_type_label} Servers" if requested_db_type else "Servers"
    ws.append(headers)
    for row in rows:
        append_excel_row(
            ws,
            [
                row["server"],
                row["ip"],
                row["environment"],
                row["db_type"],
                row["virtual"],
                row["server_model"],
                row["server_vendor"],
                row["core"],
                row["ram"],
                row["operating_system"],
            ]
        )
    apply_sheet_style(ws, headers)
    return workbook_response(workbook, filename)


@router.get("/databases.xlsx")
def export_databases(
    db_type: str | None = None,
    environment: str | None = None,
    role: str | None = None,
    db: Session = Depends(get_db),
):
    maybe_refresh_zabbix_cache(db)
    stmt = (
        select(DatabaseInstance)
        .options(selectinload(DatabaseInstance.host))
        .order_by(DatabaseInstance.db_type, DatabaseInstance.name)
    )
    stmt = apply_database_filters(stmt, db_type, environment, role)
    databases = db.scalars(stmt).all()

    headers = [
        "db_type",
        "name",
        "host",
        "environment",
        "version",
        "port",
        "service_name",
        "status",
        "powa_repository",
        "powa_server_name",
        "powa_database_name",
        "last_snapshot",
    ]
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Databases"
    ws.append(headers)
    for database in databases:
        append_excel_row(
            ws,
            [
                database.db_type,
                database.name,
                database.host.hostname,
                database.environment,
                database.version,
                database.port,
                database.service_name,
                database.status,
                database.powa_repository,
                database.powa_server_name,
                database.powa_database_name,
                database.last_snapshot,
            ]
        )
    apply_sheet_style(ws, headers)
    return workbook_response(workbook, "dba_inventory_databases.xlsx")
